from paddleocr import PaddleOCR
import os
import cv2
from openpyxl import Workbook

# Initialize PaddleOCR
ocr = PaddleOCR(use_angle_cls=True, lang='en')

# Directory containing the cropped column images
input_dir = 'preprocessed_cropped_columns'

# Lists to store extracted text and column names from all images
all_extracted_text = []
column_names = []

# Iterate through all images in the input directory
for filename in sorted(os.listdir(input_dir)):
    if filename.endswith(('.png', '.jpg', '.jpeg')):  # Add more extensions if needed
        image_path = os.path.join(input_dir, filename)

        # Read the image
        img = cv2.imread(image_path)

        # Perform OCR
        result = ocr.ocr(img, cls=True)

        # Initialize a list to store text lines for the current image
        extracted_text = []

        # Check if the result is not None and contains text
        if result is not None:
            for line in result:
                for word_info in line:
                    extracted_text.append(word_info[1][0])  # Append the recognized text

        # Only add the text if there is any extracted text
        if extracted_text:
            # Extract the first word to use as the column name
            first_word = extracted_text[0].split()[0]
            column_names.append(first_word)
            # Join all extracted text for this image, excluding the first word
            full_text = '\n'.join(extracted_text[1:]) if len(extracted_text[0].split()) == 1 else '\n'.join([' '.join(extracted_text[0].split()[1:])] + extracted_text[1:])
            # Add the extracted text to the list
            all_extracted_text.append(full_text)

            print(f"Processed {filename}")
        else:
            print(f"No text found in {filename}, skipping.")

# Create a new workbook and select the active sheet
workbook = Workbook()
sheet = workbook.active

# Write the extracted text to the Excel file
for col_num, (col_name, text) in enumerate(zip(column_names, all_extracted_text), 1):
    col_letter = sheet.cell(row=1, column=col_num).column_letter
    sheet[f"{col_letter}1"] = col_name  # Use the first word as the column name
    
    # Split the text by newlines and write each line to a separate cell
    for row_num, line in enumerate(text.split('\n'), 2):
        sheet.cell(row=row_num, column=col_num, value=line)

# Save the workbook
excel_filename = 'extracted_text.xlsx'
workbook.save(excel_filename)

print(f"Extracted text from {len(all_extracted_text)} images")
print(f"Extracted text has been saved to '{excel_filename}'")
